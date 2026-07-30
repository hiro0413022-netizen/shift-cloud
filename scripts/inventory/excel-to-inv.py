# -*- coding: utf-8 -*-
"""202606_ゴルフウィング在庫リスト.xlsm → inv_* シード（DECISIONS #96）

エクセルの横持ち（1ヶ月＝確認日/時間/担当/数量の4列）を縦持ちに変換する。
棚卸は前月比で30〜90品番しか動かないため、
  ・最古の 2025-12 は全量
  ・以降は「前月をコピー → 消えた品番を削除 → 変わった数量だけUPDATE」
という差分方式で書き出す（数値は原本と完全一致。生成後に検証クエリで突合する）。
担当・時刻はセッション内の最頻値を既定とし、それ以外の品番だけ明示的に上書きする。
"""
import openpyxl, datetime, os
from collections import defaultdict, Counter
from openpyxl.utils.datetime import from_excel

SRC = '/sessions/eloquent-gallant-knuth/mnt/uploads/202606_ゴルフウィング在庫リスト.xlsm'
OUT = '/sessions/eloquent-gallant-knuth/mnt/YOZAN GENESIS/scripts/inventory'
C = 'ec00ad2a-4032-4061-bdb7-03face8a04e7'
S = '82bb4e18-427d-4cc7-a834-c9e2a9b18199'
GROUPS = [34, 30, 26, 22, 18, 14, 10]        # 古い月 → 新しい月

def q(v):
    if v is None: return 'null'
    # Excelの浮動小数ノイズ（985.9999999999999 等）を落とす。金額なので小数2桁で十分
    if isinstance(v, float):
        r = round(v, 2)
        return repr(int(r)) if r == int(r) else repr(r)
    if isinstance(v, int): return repr(v)
    s = str(v).strip()
    return 'null' if s == '' else "'" + s.replace("'", "''") + "'"

def qt(v):
    """テキスト列用。原本の「仕様」「カラー、仕様」欄に数値が入っている行があるため必ず文字列化する"""
    if v is None: return 'null'
    if isinstance(v, float) and v == int(v): v = int(v)
    return q(str(v))

def dt(v):
    if isinstance(v, (int, float)):
        x = from_excel(v); return x.date() if isinstance(x, datetime.datetime) else x
    if isinstance(v, datetime.datetime): return v.date()
    return v if isinstance(v, datetime.date) else None

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['ゴルフウィング在庫リスト']

items, snap, meta, sess = [], {}, {}, {}
for g in GROUPS:
    sess[g] = dt(ws.cell(2, g + 3).value); snap[g] = {}; meta[g] = {}
for r in range(3, ws.max_row + 1):
    code = ws.cell(r, 1).value
    if not code: continue
    code = str(code).strip()
    items.append([code] + [ws.cell(r, c).value for c in (2,3,4,5,6,7,8,9,38,40,41)])
    for g in GROUPS:
        v = ws.cell(r, g + 3).value
        if not isinstance(v, (int, float)): continue
        snap[g][code] = int(v)
        d = dt(ws.cell(r, g).value) or sess[g]
        t = ws.cell(r, g + 1).value
        who = ws.cell(r, g + 2).value
        meta[g][code] = ('%s %s+09' % (d, t.strftime('%H:%M:%S') if isinstance(t, datetime.time) else '12:00:00'),
                         str(who).strip() if who else None)

def sess_sql(g):
    on = sess[g]
    return ("insert into inv_count_sessions (company_id,store_id,counted_on,label,status,closed_at,note) values "
            "(%s,%s,%s,%s,'closed',%s,%s) on conflict do nothing;" % (
        q(C), q(S), q(str(on)), q('%d年%d月末 棚卸' % (on.year, on.month)), q('%s 23:59:59+09' % on),
        q('エクセル(202606_ゴルフウィング在庫リスト.xlsm)からの移行分。当時は入出庫台帳が無く理論在庫が存在しないため theoretical/diff は null')))

def sid(g):
    return "(select id from inv_count_sessions where company_id=%s and counted_on=%s and deleted_at is null)" % (q(C), q(str(sess[g])))

L = ['-- 202606_ゴルフウィング在庫リスト.xlsm からの移行（DECISIONS #96 / 生成: scripts/inventory/excel-to-inv.py）',
     '-- 数量は原本と完全一致。再実行しても増えない（on conflict / 差分UPDATE）', '']

# ---- 品番マスタ ----
# 管理番号は CAT-MK-連番 なので category/maker は inv_codes への join で復元できる
# （原本362行すべてで一致することを確認済み）。単位・保管場所は出現数が少ないので番号で持つ。
UNITS = ['個', '本', '枚', 'ダース', '箱']
LOCS  = sorted({str(it[7]).strip() for it in items if it[7]})
L.append("insert into inv_items (company_id,store_id,code,category,maker,name,spec,variant,unit,location1,location2,notes,list_price,cost_price)")
L.append("select %s,%s,v.code,cat.name,mk.name,v.name,v.spec,v.variant," % (q(C), q(S)))
L.append("  (array[%s])[v.u], case when v.l>0 then (array[%s])[v.l] end, v.location2, v.notes, v.list_price, v.cost_price"
         % (",".join(q(u) for u in UNITS), ",".join(q(x) for x in LOCS)))
L.append("from (values")
L.append(",\n".join("(%s,%s,%s,%s,%d,%d,%s,%s,%s,%s)" % (
        q(it[0]), q(it[3]), qt(it[4]), qt(it[5]),
        UNITS.index(str(it[6]).strip()) + 1 if it[6] else 1,
        LOCS.index(str(it[7]).strip()) + 1 if it[7] else 0,
        qt(it[8]), qt(it[9]), q(it[10]), q(it[11]))
    for it in items))
L.append(") v(code,name,spec,variant,u,l,location2,notes,list_price,cost_price)")
L.append("join inv_codes cat on cat.company_id=%s and cat.kind='category' and cat.abbr=split_part(v.code,'-',1) and cat.deleted_at is null" % q(C))
L.append("join inv_codes mk  on mk.company_id=%s  and mk.kind='maker'    and mk.abbr=split_part(v.code,'-',2) and mk.deleted_at  is null" % q(C))
L.append("on conflict do nothing;")
L.append("")

# ---- 棚卸: 最古は全量 ----
base = GROUPS[0]
L.append('-- %s 棚卸（%d品番・全量）' % (sess[base], len(snap[base])))
L.append(sess_sql(base))
buckets = defaultdict(list)
for c_, v in snap[base].items(): buckets[meta[base][c_]].append((c_, v))
for (at, who), pairs in buckets.items():
    L.append("insert into inv_counts (company_id,session_id,item_id,qty,counted_at,counted_by_name)")
    L.append("select %s,%s,i.id,v.q,%s::timestamptz,%s from (values" % (q(C), sid(base), q(at), q(who)))
    L.append(",".join("(%s,%d)" % (q(c_), n) for c_, n in pairs))
    L.append(") v(c,q) join inv_items i on i.company_id=%s and i.code=v.c and i.deleted_at is null" % q(C))
    L.append("on conflict (session_id,item_id) do update set qty=excluded.qty;")
L.append("")

# ---- 以降は差分 ----
for prev, g in zip(GROUPS, GROUPS[1:]):
    on = sess[g]
    cur, old = snap[g], snap[prev]
    gone    = sorted(set(old) - set(cur))
    added   = sorted(set(cur) - set(old))          # 前月に無かった品番＝UPDATEでは入らないのでINSERTする
    changed = sorted(c_ for c_ in set(cur) & set(old) if old[c_] != cur[c_])
    top = Counter(meta[g].values()).most_common(1)[0][0]
    over = defaultdict(list)
    for c_, m in meta[g].items():
        if m != top: over[m].append(c_)
    L.append('-- %s 棚卸（%d品番 / 前月から 消滅%d・変化%d）' % (on, len(cur), len(gone), len(changed)))
    L.append(sess_sql(g))
    L.append("insert into inv_counts (company_id,session_id,item_id,qty,counted_at,counted_by_name)")
    L.append("select %s,%s,c.item_id,c.qty,%s::timestamptz,%s from inv_counts c where c.session_id=%s"
             % (q(C), sid(g), q(top[0]), q(top[1]), sid(prev)))
    L.append("on conflict (session_id,item_id) do nothing;")
    if gone:
        L.append("delete from inv_counts where session_id=%s and item_id in (select id from inv_items where company_id=%s and code in (%s));"
                 % (sid(g), q(C), ",".join(q(x) for x in gone)))
    if added:
        L.append("insert into inv_counts (company_id,session_id,item_id,qty,counted_at,counted_by_name)")
        L.append("select %s,%s,i.id,v.q,%s::timestamptz,%s from (values %s) v(c,q)"
                 % (q(C), sid(g), q(top[0]), q(top[1]), ",".join("(%s,%d)" % (q(c_), cur[c_]) for c_ in added)))
        L.append("join inv_items i on i.company_id=%s and i.code=v.c and i.deleted_at is null" % q(C))
        L.append("on conflict (session_id,item_id) do update set qty=excluded.qty;")
    if changed:
        L.append("update inv_counts t set qty=v.q from (values %s) v(c,q)"
                 % ",".join("(%s,%d)" % (q(c_), cur[c_]) for c_ in changed))
        L.append("join inv_items i on i.company_id=%s and i.code=v.c where t.session_id=%s and t.item_id=i.id;" % (q(C), sid(g)))
    for (at, who), cds in over.items():
        L.append("update inv_counts t set counted_at=%s::timestamptz, counted_by_name=%s from inv_items i" % (q(at), q(who)))
        L.append("where i.company_id=%s and i.code in (%s) and t.session_id=%s and t.item_id=i.id;"
                 % (q(C), ",".join(q(x) for x in cds), sid(g)))
    L.append("")

L.append("""-- 集計値の確定（inv_close_count は通さない＝過去分に adjust を起票しない）
update inv_count_sessions s set total_qty=agg.qty, total_value=agg.val, updated_at=now()
from (select c.session_id, sum(c.qty) qty, sum(c.qty*coalesce(i.cost_price,0)) val
      from inv_counts c join inv_items i on i.id=c.item_id group by c.session_id) agg
where agg.session_id=s.id and s.company_id=%s;""" % q(C))

L.append("")
L.append("-- 検証: 原本の 品番数 / 合計数量 と一致すること")
L.append("-- " + " | ".join("%s:%d件%d点" % (sess[g], len(snap[g]), sum(snap[g].values())) for g in GROUPS))

sql = "\n".join(L) + "\n"
os.makedirs(OUT, exist_ok=True)
open(OUT + '/seed_202606.sql', 'w', encoding='utf-8').write(sql)

# 検証用: 原本の正解値
exp = []
for g in GROUPS:
    exp.append((str(sess[g]), len(snap[g]), sum(snap[g].values())))
open('/tmp/expected.txt','w').write("\n".join("%s\t%d\t%d" % e for e in exp))

os.makedirs('/tmp/c3', exist_ok=True)
blocks, cur_b = [], []
for line in L:
    cur_b.append(line)
    if line.rstrip().endswith(';'):
        blocks.append("\n".join(cur_b)); cur_b = []
for i, b in enumerate(blocks):
    open('/tmp/c3/%02d.sql' % i, 'w', encoding='utf-8').write(b)
print('blocks', len(blocks), 'total', len(sql))
for i,b in enumerate(blocks): print(' ', i, len(b))
print(open('/tmp/expected.txt').read())
