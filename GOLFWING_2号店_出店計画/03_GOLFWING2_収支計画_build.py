#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GOLF WING 2号店 収支計画モデル（非居抜き・3〜5打席・相互利用・8月値上げ後料金）
- 打席数/坪数を入力すると 会員上限・内装CAPEX・機材CAPEX・家賃が自動連動
- 11シート / 楽観・標準・悲観・最悪 / 36ヶ月 / 全数式連動 / 前受けON-OFF / 防犯は開業N ヶ月後
実行: python 03_GOLFWING2_収支計画_build.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
import os

YEN='#,##0;(#,##0);"-"'; PCT='0.0%'; BASE='Arial'
f_title=Font(name=BASE,bold=True,size=13,color='1F3864')
f_h=Font(name=BASE,bold=True,color='FFFFFF'); f_sub=Font(name=BASE,bold=True,color='1F3864')
f_in=Font(name=BASE,color='0000FF'); f_calc=Font(name=BASE,color='000000'); f_link=Font(name=BASE,color='008000')
f_note=Font(name=BASE,italic=True,size=9,color='808080')
fill_in=PatternFill('solid',fgColor='FFFF00'); fill_h=PatternFill('solid',fgColor='1F3864')
fill_sub=PatternFill('solid',fgColor='D9E1F2'); fill_warn=PatternFill('solid',fgColor='FFC7CE')
thin=Side(style='thin',color='BFBFBF'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal='center',vertical='center',wrap_text=True)
left=Alignment(horizontal='left',vertical='center',wrap_text=True)
wb=Workbook()
def hdr(ws,row,headers,start=1):
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=start+i,value=h); c.font=f_h; c.fill=fill_h; c.alignment=center; c.border=border
def inp(ws,coord,val,fmt=YEN):
    c=ws[coord]; c.value=val; c.font=f_in; c.fill=fill_in; c.number_format=fmt; c.border=border; return c
def calc(ws,coord,formula,fmt=YEN,link=False):
    c=ws[coord]; c.value=formula; c.font=f_link if link else f_calc; c.number_format=fmt; c.border=border; return c
def label(ws,coord,text,sub=False):
    c=ws[coord]; c.value=text; c.font=f_sub if sub else Font(name=BASE)
    if sub: c.fill=fill_sub
    c.alignment=left; return c

# ===== Scenarios =====
ws=wb.active; ws.title='Scenarios'
ws['A1']='シナリオ設定（4パターン）'; ws['A1'].font=f_title
ws['A2']='選択シナリオ →'; ws['A2'].font=f_sub
inp(ws,'B2','標準',fmt='General')
dv=DataValidation(type='list',formula1='"楽観,標準,悲観,最悪"',allow_blank=False); ws.add_data_validation(dv); dv.add(ws['B2'])
ws['D2']='B2で全シート自動更新'; ws['D2'].font=f_note
hdr(ws,4,['パラメータ','楽観','標準','悲観','最悪','選択中(自動)','備考'])
params=[
 ('平均月会費(税抜)',18500,17000,15500,14000,YEN,'8月値上げ後(レギュラー19,800/マスター24,800等)の加重平均を保守補正'),
 ('入会率(体験→入会)',0.65,0.55,0.35,0.25,PCT,'宝塚の高転換を参考に保守設定'),
 ('退会率(月次)',0.03,0.04,0.06,0.08,PCT,'相互利用で定着改善を見込むが保守'),
 ('月間体験数 Y1',22,16,10,6,'#,##0','初年度'),
 ('月間体験数 Y2',30,22,14,9,'#,##0',''),
 ('月間体験数 Y3',36,26,18,11,'#,##0',''),
 ('広告費 開業1-3ヶ月(月)',150000,180000,220000,280000,YEN,'大型店の開業集客'),
 ('広告費 4ヶ月目以降(月)',80000,100000,130000,170000,YEN,''),
 ('人件費(月)',700000,850000,1000000,1200000,YEN,'3-5打席・一部有人。最悪は増員'),
 ('物販ARPU(会員/月)',3500,2500,1800,1200,YEN,''),
 ('開業時先行会員数',40,30,12,6,'#,##0','宝塚会員の相互利用移行含む'),
 ('追加CAPEX(初期不良等)',0,0,500000,1200000,YEN,'悲観/最悪'),
]
r0=5
for i,(lb,o,s,p,w,fmt,note) in enumerate(params):
    r=r0+i; label(ws,f'A{r}',lb)
    inp(ws,f'B{r}',o,fmt); inp(ws,f'C{r}',s,fmt); inp(ws,f'D{r}',p,fmt); inp(ws,f'E{r}',w,fmt)
    calc(ws,f'F{r}',f'=INDEX(B{r}:E{r},MATCH($B$2,$B$4:$E$4,0))',fmt)
    ws[f'G{r}']=note; ws[f'G{r}'].font=f_note
for col,wd in zip('ABCDEFG',[24,11,11,11,11,13,40]): ws.column_dimensions[col].width=wd
SF={'fee':'F5','join':'F6','churn':'F7','tY1':'F8','tY2':'F9','tY3':'F10','adO':'F11','adA':'F12','labor':'F13','arpu':'F14','pre':'F15','addcapex':'F16'}

# ===== Assumptions =====
ws=wb.create_sheet('Assumptions')
ws['A1']='前提条件（入力の集約）'; ws['A1'].font=f_title
ws['A2']='黄=入力／緑=参照。打席数・坪数を変えると会員上限・内装・機材・家賃が自動連動'; ws['A2'].font=f_note
label(ws,'A4','■ 店舗規模・基本前提',sub=True)
rows_const=[
 ('打席数',4,'#,##0'),               # B5
 ('坪数',48,'#,##0'),                # B6
 ('会員/打席(上限係数)',45,'#,##0'), # B7
 ('入会金(税抜)',30000,YEN),         # B8
 ('体験単価(税抜)',2200,YEN),        # B9
 ('物販原価率',0.65,PCT),            # B10
 ('決済手数料率',0.03,PCT),          # B11
 ('決済適用率',0.90,PCT),            # B12
 ('減価償却年数(設備)',8,'#,##0'),    # B13
 ('実効税率',0.30,PCT),              # B14
 ('賃料坪単価(月)',12000,YEN),       # B15
 ('内装坪単価(非居抜き)',300000,YEN),# B16
 ('シミュレーター単価/打席',1300000,YEN), # B17
 ('光熱費(月)',120000,YEN),          # B18
 ('通信費(月)',15000,YEN),           # B19
 ('保険(月)',15000,YEN),             # B20
 ('システム費(月)',10000,YEN),       # B21 自社Genesis
 ('清掃費(月)',30000,YEN),           # B22
 ('修繕・設備更新積立(月)',40000,YEN),# B23
 ('消耗品費(月)',20000,YEN),         # B24
 ('税理士費用(月)',0,YEN),           # B25 YOZAN税理士
 ('予備費率(対その他固定費)',0.03,PCT),# B26
 ('保証金月数',6,'#,##0'),           # B27
 ('出資金',20000000,YEN),            # B28
 ('自己資金・融資',0,YEN),           # B29
]
r=5
for lb,v,fmt in rows_const:
    label(ws,f'A{r}',lb); inp(ws,f'B{r}',v,fmt); r+=1
# derived
label(ws,'A31','会員上限(=打席×係数)'); calc(ws,'B31','=B5*B7','#,##0')
label(ws,'A32','家賃(=坪数×賃料坪単価)'); calc(ws,'B32','=B6*B15',YEN)
label(ws,'A33','■ 選択シナリオ反映値（自動）',sub=True)
active=[('平均月会費(税抜)',SF['fee'],YEN),('入会率',SF['join'],PCT),('退会率(月)',SF['churn'],PCT),
 ('月間体験数 Y1',SF['tY1'],'#,##0'),('月間体験数 Y2',SF['tY2'],'#,##0'),('月間体験数 Y3',SF['tY3'],'#,##0'),
 ('広告費 開業1-3ヶ月(月)',SF['adO'],YEN),('広告費 4ヶ月目以降(月)',SF['adA'],YEN),('人件費(月)',SF['labor'],YEN),
 ('物販ARPU(月)',SF['arpu'],YEN),('開業時先行会員数',SF['pre'],'#,##0'),('追加CAPEX',SF['addcapex'],YEN)]
r=34
for lb,ref,fmt in active:
    label(ws,f'A{r}',lb); calc(ws,f'B{r}',f'=Scenarios!{ref}',fmt,link=True); r+=1
# r=46
label(ws,'A47','その他固定費小計'); calc(ws,'B47','=B18+B19+B20+B21+B22+B23+B24+B25',YEN)
label(ws,'A48','予備費'); calc(ws,'B48','=B47*B26',YEN)
label(ws,'A49','その他固定費 計'); calc(ws,'B49','=B47+B48',YEN)
label(ws,'A51','■ 前受け（月会費先取り）',sub=True)
label(ws,'A52','前受けON(1=有効,0=無効)'); inp(ws,'B52',1,'#,##0')
label(ws,'A53','前受けフロート係数(月分)'); inp(ws,'B53',1.5,'0.0')
label(ws,'A54','■ 後日投資（防犯・セキュリティ）',sub=True)
label(ws,'A55','防犯・セキュリティ投資額'); inp(ws,'B55',500000,YEN)
label(ws,'A56','実施月(開業何ヶ月後)'); inp(ws,'B56',6,'#,##0')
ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=15
A={'cap':'Assumptions!$B$31','join_fee':'Assumptions!$B$8','trial_p':'Assumptions!$B$9','cogs':'Assumptions!$B$10',
 'pay':'Assumptions!$B$11','payr':'Assumptions!$B$12','dep_y':'Assumptions!$B$13','tax':'Assumptions!$B$14',
 'rent':'Assumptions!$B$32','fee':'Assumptions!$B$34','join':'Assumptions!$B$35','churn':'Assumptions!$B$36',
 'tY1':'Assumptions!$B$37','tY2':'Assumptions!$B$38','tY3':'Assumptions!$B$39','adO':'Assumptions!$B$40',
 'adA':'Assumptions!$B$41','labor':'Assumptions!$B$42','arpu':'Assumptions!$B$43','pre':'Assumptions!$B$44',
 'addcapex':'Assumptions!$B$45','otherfix':'Assumptions!$B$49','secAmt':'Assumptions!$B$55','secMon':'Assumptions!$B$56',
 'ppON':'Assumptions!$B$52','ppCoef':'Assumptions!$B$53'}

# ===== CAPEX =====
ws=wb.create_sheet('CAPEX')
ws['A1']='初期投資（CAPEX）内訳 ※非居抜き'; ws['A1'].font=f_title
ws['A2']='内装=坪数×内装坪単価、機材=打席数×単価/打席 で自動連動。防犯は初期投資外(開業N ヶ月後)'; ws['A2'].font=f_note
hdr(ws,4,['項目','金額','備考'])
capex=[
 ('内装工事(非居抜き)','=Assumptions!$B$6*Assumptions!$B$16','坪数×坪単価。スケルトンから造作・打席ブース'),
 ('電気工事',1200000,'分電盤・打席電源・照明'),
 ('空調・換気・設備工事',1500000,'大型・打席発熱対応'),
 ('防音・安全対策',1200000,'防音・防球・防護'),
 ('シミュレーター・機材','=Assumptions!$B$5*Assumptions!$B$17','打席数×単価/打席'),
 ('PC・ネットワーク・AV',700000,''),
 ('什器・備品',700000,''),
 ('看板・サイン',500000,''),
 ('決済・POS環境',200000,''),
 ('HP制作・初期ブランディング',300000,'YOZAN WEB'),
 ('開業前費用(採用・研修・内覧)',400000,''),
]
r=5
for lb,v,note in capex:
    label(ws,f'A{r}',lb)
    if isinstance(v,str): calc(ws,f'B{r}',v,YEN,link=True)
    else: inp(ws,f'B{r}',v)
    ws[f'C{r}']=note; ws[f'C{r}'].font=f_note; r+=1
last=r-1  # 15
label(ws,f'A{r}','設備小計',sub=True); calc(ws,f'B{r}',f'=SUM(B5:B{last})'); SUB=r; r+=1
label(ws,f'A{r}','予備費(10%)',sub=True); calc(ws,f'B{r}',f'=B{SUB}*0.1'); PRE=r; r+=1
label(ws,f'A{r}','CAPEX計(減価償却対象)',sub=True); calc(ws,f'B{r}',f'=B{SUB}+B{PRE}'); DEPB=r; r+=1
label(ws,f'A{r}','保証金・敷金',sub=True); calc(ws,f'B{r}','=Assumptions!$B$32*Assumptions!$B$27',YEN,link=True); DEPO=r; r+=1
label(ws,f'A{r}','追加CAPEX(シナリオ)',sub=True); calc(ws,f'B{r}',f'={A["addcapex"]}',YEN,link=True); ADD=r; r+=1
label(ws,f'A{r}','総初期投資',sub=True); calc(ws,f'B{r}',f'=B{DEPB}+B{DEPO}+B{ADD}'); TOT=r
ws[f'B{TOT}'].font=Font(name=BASE,bold=True,color='C00000')
CAPEX_DEPBASE=f'CAPEX!$B${DEPB}'; CAPEX_ADD=f'CAPEX!$B${ADD}'; CAPEX_TOTAL=f'CAPEX!$B${TOT}'
ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=14; ws.column_dimensions['C'].width=36

# ===== OPEX =====
ws=wb.create_sheet('OPEX')
ws['A1']='月次コスト内訳（OPEX）'; ws['A1'].font=f_title
hdr(ws,3,['区分','項目','金額(月)','種別','備考'])
opex=[
 ('固定費','家賃(坪数×坪単価)','=Assumptions!$B$32','固定',''),
 ('固定費','人件費','=Assumptions!$B$42','固定(シナリオ)',''),
 ('固定費','広告費(開業1-3ヶ月)','=Assumptions!$B$40','固定(期間変動)','開業直後'),
 ('固定費','広告費(4ヶ月〜)','=Assumptions!$B$41','固定(期間変動)','4ヶ月目以降'),
 ('固定費','光熱費','=Assumptions!$B$18','固定',''),
 ('固定費','通信費','=Assumptions!$B$19','固定',''),
 ('固定費','保険','=Assumptions!$B$20','固定',''),
 ('固定費','システム費','=Assumptions!$B$21','固定','自社Genesis'),
 ('固定費','清掃費','=Assumptions!$B$22','固定',''),
 ('固定費','修繕・設備更新積立','=Assumptions!$B$23','固定',''),
 ('固定費','消耗品費','=Assumptions!$B$24','固定',''),
 ('固定費','税理士費用','=Assumptions!$B$25','固定','YOZAN税理士=0'),
 ('固定費','予備費','=Assumptions!$B$48','固定',''),
 ('変動費','決済手数料','売上×3%×適用率','変動','Monthly_PL'),
 ('変動費','物販原価','物販売上×65%','変動','Monthly_PL'),
]
r=4
for seg,it,val,typ,note in opex:
    ws[f'A{r}']=seg; ws[f'B{r}']=it
    if isinstance(val,str) and val.startswith('='): calc(ws,f'C{r}',val,YEN,link=True)
    else: ws[f'C{r}']=val; ws[f'C{r}'].font=f_note
    ws[f'D{r}']=typ; ws[f'E{r}']=note; ws[f'E{r}'].font=f_note
    for col in 'ABDE': ws[f'{col}{r}'].border=border
    r+=1
for col,wd in zip('ABCDE',[10,22,15,14,22]): ws.column_dimensions[col].width=wd

# ===== Monthly_PL =====
ws=wb.create_sheet('Monthly_PL')
ws['A1']='月次損益計画（36ヶ月）'; ws['A1'].font=f_title
cols=['月','年','期首会員','体験数','入会数','退会数','期末会員','月会費売上','入会金売上','体験売上','物販売上','売上合計','物販原価','決済手数料','変動費計','家賃','人件費','広告費','その他固定費','固定費計','減価償却','営業利益','税前利益','法人税等','税引後利益']
hdr(ws,3,cols); HR=3
for m in range(1,37):
    r=HR+m
    ws[f'A{r}']=m; ws[f'B{r}']=f'=ROUNDUP(A{r}/12,0)'
    ws[f'C{r}']=(f'={A["pre"]}' if m==1 else f'=G{r-1}')
    ws[f'D{r}']=f'=ROUND(IF(B{r}=1,{A["tY1"]},IF(B{r}=2,{A["tY2"]},{A["tY3"]}))*MIN(1,A{r}/6),0)'
    ws[f'F{r}']=f'=ROUND(C{r}*{A["churn"]},0)'
    ws[f'E{r}']=f'=MAX(0,MIN(ROUND(D{r}*{A["join"]},0),{A["cap"]}-(C{r}-F{r})))'
    ws[f'G{r}']=f'=MAX(0,MIN({A["cap"]},C{r}+E{r}-F{r}))'
    ws[f'H{r}']=f'=C{r}*{A["fee"]}'; ws[f'I{r}']=f'=E{r}*{A["join_fee"]}'; ws[f'J{r}']=f'=D{r}*{A["trial_p"]}'; ws[f'K{r}']=f'=C{r}*{A["arpu"]}'
    ws[f'L{r}']=f'=SUM(H{r}:K{r})'
    ws[f'M{r}']=f'=K{r}*{A["cogs"]}'; ws[f'N{r}']=f'=L{r}*{A["pay"]}*{A["payr"]}'; ws[f'O{r}']=f'=M{r}+N{r}'
    ws[f'P{r}']=f'={A["rent"]}'; ws[f'Q{r}']=f'={A["labor"]}'; ws[f'R{r}']=f'=IF(A{r}<=3,{A["adO"]},{A["adA"]})'; ws[f'S{r}']=f'={A["otherfix"]}'
    ws[f'T{r}']=f'=P{r}+Q{r}+R{r}+S{r}'
    ws[f'U{r}']=f'=({CAPEX_DEPBASE}+{CAPEX_ADD})/({A["dep_y"]}*12)+IF(A{r}>={A["secMon"]},{A["secAmt"]}/({A["dep_y"]}*12),0)'
    ws[f'V{r}']=f'=L{r}-O{r}-T{r}-U{r}'; ws[f'W{r}']=f'=V{r}'; ws[f'X{r}']=f'=MAX(0,W{r})*{A["tax"]}'; ws[f'Y{r}']=f'=W{r}-X{r}'
    for ci in range(1,26):
        c=ws.cell(row=r,column=ci); c.border=border
        c.number_format='#,##0' if c.column_letter in ('A','B','C','D','E','F','G') else YEN
        c.font=f_link if (isinstance(c.value,str) and ('Assumptions!' in c.value or 'CAPEX!' in c.value)) else f_calc
for col in 'ABCDEFG': ws.column_dimensions[col].width=8
for col in ['H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y']: ws.column_dimensions[col].width=11
ws.freeze_panes='C4'
PL_V='Monthly_PL!$V'; PL_Y='Monthly_PL!$Y'; PL_U='Monthly_PL!$U'

# ===== Cash_Flow =====
ws=wb.create_sheet('Cash_Flow')
ws['A1']='月次キャッシュフロー'; ws['A1'].font=f_title
ws['A2']='期末現金マイナス=資金ショート(赤)。営業CF=税引後+減価償却。前受け=運転資本。'; ws['A2'].font=f_note
hdr(ws,3,['月','営業CF','投資CF','財務CF','前受けCF(運転資本)','純増減','期末現金','累計営業CF','回収判定(内部)'])
ws['A4']=0; ws['B4']=0; ws['C4']=f'=-{CAPEX_TOTAL}'; ws['C4'].font=f_link
ws['D4']='=Assumptions!$B$28+Assumptions!$B$29'; ws['D4'].font=f_link
ws['E4']=0; ws['F4']='=B4+C4+D4+E4'; ws['G4']='=F4'; ws['H4']=0; ws['I4']=9999
for col in 'ABCDEFGHI': ws[f'{col}4'].border=border; ws[f'{col}4'].number_format=YEN
for m in range(1,37):
    r=4+m; plr=3+m; prev='0' if m==1 else f'Monthly_PL!$G${plr-1}'
    ws[f'A{r}']=m
    ws[f'B{r}']=f'={PL_Y}{plr}+{PL_U}{plr}'; ws[f'B{r}'].font=f_link
    ws[f'C{r}']=f'=IF(A{r}={A["secMon"]},-{A["secAmt"]},0)'; ws[f'C{r}'].font=f_link
    ws[f'D{r}']=0
    ws[f'E{r}']=f'={A["ppON"]}*{A["ppCoef"]}*{A["fee"]}*(Monthly_PL!$G${plr}-{prev})'; ws[f'E{r}'].font=f_link
    ws[f'F{r}']=f'=B{r}+C{r}+D{r}+E{r}'; ws[f'G{r}']=f'=G{r-1}+F{r}'; ws[f'H{r}']=f'=SUM($B$5:B{r})'
    ws[f'I{r}']=f'=IF(H{r}>={CAPEX_TOTAL},A{r},9999)'
    for col in 'ABCDEFGHI': ws[f'{col}{r}'].border=border; ws[f'{col}{r}'].number_format=YEN
    ws[f'A{r}'].number_format='#,##0'
ws.conditional_formatting.add('G5:G40',CellIsRule(operator='lessThan',formula=['0'],fill=fill_warn,font=Font(name=BASE,color='9C0006')))
for col,wd in zip('ABCDEFGHI',[8,13,11,11,15,13,14,14,12]): ws.column_dimensions[col].width=wd
ws.freeze_panes='B4'

# ===== BreakEven =====
ws=wb.create_sheet('BreakEven')
ws['A1']='損益分岐点分析'; ws['A1'].font=f_title
rows=[
 ('月固定費(定常/4ヶ月〜)','=Assumptions!$B$32+Assumptions!$B$42+Assumptions!$B$41+Assumptions!$B$49',YEN,True),
 ('平均月会費(税抜)','=Assumptions!$B$34',YEN,True),
 ('物販ARPU','=Assumptions!$B$43',YEN,True),
 ('物販原価率','=Assumptions!$B$10',PCT,True),
 ('決済実効率','=Assumptions!$B$11*Assumptions!$B$12',PCT,True),
 ('会員あたり月次貢献','=B3*(1-B6)+B4*(1-B5)-B4*B6',YEN,False),
 ('損益分岐点 会員数','=ROUNDUP(B2/B7,0)','#,##0',False),
 ('会員上限(打席×係数)','=Assumptions!$B$31','#,##0',True),
 ('必要月商(概算)','=B8*(B3+B4)',YEN,False),
 ('現状会員(36ヶ月)','=Monthly_PL!$G$39','#,##0',True),
 ('BEP維持に必要な月間入会','=ROUNDUP(B8*Assumptions!$B$36,0)','#,##0',False),
 ('必要月間体験数','=ROUNDUP(B12/Assumptions!$B$35,0)','#,##0',False),
]
r=2
for lb,val,fmt,link in rows:
    label(ws,f'A{r}',lb); calc(ws,f'B{r}',val,fmt,link=link); r+=1
ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=14

# ===== Sensitivity =====
ws=wb.create_sheet('Sensitivity')
ws['A1']='感応度分析（年間営業利益への影響・一次近似）'; ws['A1'].font=f_title
ws['A2']='精緻な再計算はAssumptions変更→Summaryで確認。下表は目安'; ws['A2'].font=f_note
label(ws,'A4','基準:年間営業利益(Y3)',sub=True); calc(ws,'B4','=SUM(Monthly_PL!$V$28:$V$39)',YEN,link=True)
hdr(ws,6,['変動要因','変化幅','年間営業利益への影響(近似)'])
sens=[('打席数','+1打席(会員上限+係数)','=Assumptions!$B$7*BreakEven!$B$7*12'),
 ('会員数','+10%','=0.1*Monthly_PL!$G$39*BreakEven!$B$7*12'),
 ('平均月会費','+1,000円','=Monthly_PL!$G$39*1000*(1-Assumptions!$B$11*Assumptions!$B$12)*12'),
 ('退会率','+2pt','=-Monthly_PL!$G$39*0.02*BreakEven!$B$7*12'),
 ('人件費','+10%','=-Assumptions!$B$42*0.1*12'),
 ('家賃','+10%','=-Assumptions!$B$32*0.1*12'),
 ('内装坪単価','+5万/坪','=-Assumptions!$B$6*50000/(Assumptions!$B$13*12)*12')]
r=7
for f_,d_,v_ in sens:
    ws[f'A{r}']=f_; ws[f'B{r}']=d_; calc(ws,f'C{r}',v_,YEN,link=True)
    for col in 'AB': ws[f'{col}{r}'].border=border
    r+=1
for col,wd in zip('ABC',[16,18,28]): ws.column_dimensions[col].width=wd

# ===== Risks =====
ws=wb.create_sheet('Risks')
ws['A1']='リスク一覧'; ws['A1'].font=f_title
hdr(ws,3,['分類','リスク','影響度','発生可能性','対策'])
risks=[
 ('出店','駐車場(打席×2台)確保が困難','高','高','探索条件緩和・近隣コインP併用・打席数調整'),
 ('資金','非居抜き内装で初期投資が膨らむ','高','高','内装坪単価精査・打席数縮小・段階投資・融資増額'),
 ('資金','融資が下りない/条件悪い','高','中','複数行並行・公庫＋民間・自己資金比率UP'),
 ('集客','初年度立ち上がり遅延','高','中','相互利用告知・紹介・体験導線・広告前倒し'),
 ('運営','退会率上昇','高','中','予約改善・レッスン価値・相互利用の利便性'),
 ('設備','機材納期(3-5打席分)','中','中','早期発注・段階開業・リース'),
 ('法務','ブランド/相互利用条件の相違','高','中','ファインと書面合意'),
 ('近隣','騒音クレーム','中','低','事前騒音測定・防音余裕設計'),
 ('ブランド','品質ばらつきでGOLF WING毀損','高','中','SOP・KPI監視・是正フロー'),
]
r=4
for row in risks:
    for i,v in enumerate(row):
        c=ws.cell(row=r,column=1+i,value=v); c.border=border; c.alignment=left
    r+=1
for col,wd in zip('ABCDE',[10,30,8,10,40]): ws.column_dimensions[col].width=wd

# ===== Investor_QA =====
ws=wb.create_sheet('Investor_QA')
ws['A1']='想定問答'; ws['A1'].font=f_title
hdr(ws,3,['想定質問','回答案'])
qa=[
 ('なぜ非居抜きで高い投資を?','宝塚10km圏の相互利用で会員利便と定着を高める立地優先。内装は坪単価を精査し打席数で調整可能。'),
 ('必要資金と調達は?','打席数・坪数で総初期投資が自動算出。出資＋融資＋自己資金で構成し、融資内諾を契約の前提とする。'),
 ('売上前提は妥当か?','8月値上げ後料金×会員構成。入会率・退会率は宝塚実績を保守補正。悲観/最悪で耐性確認。'),
 ('相互利用の効果は?','宝塚会員の一部が移行・併用し初速を確保。ただし会費/KPI帰属はファインと事前合意。'),
 ('打席数は何面が最適?','打席数を変えると必要資金・会員上限・損益分岐点が自動更新。3/4/5で比較し投資対効果で決定。'),
 ('資金ショートは?','Cash_Flowで月末現金を可視化、赤表示。前受けで運転資本を改善。'),
]
r=4
for q,a in qa:
    ws.cell(row=r,column=1,value=q).border=border; ws.cell(row=r,column=1).alignment=left
    ws.cell(row=r,column=2,value=a).border=border; ws.cell(row=r,column=2).alignment=left
    r+=1
ws.column_dimensions['A'].width=28; ws.column_dimensions['B'].width=72

# ===== Summary =====
ws=wb.create_sheet('Summary',0)
ws['A1']='GOLF WING 2号店 収支サマリー'; ws['A1'].font=f_title
ws['A2']='選択シナリオ:'; ws['A2'].font=f_sub
calc(ws,'B2','=Scenarios!$B$2','General',link=True); ws['B2'].font=Font(name=BASE,bold=True,color='C00000')
ws['D2']='打席数・坪数はAssumptionsで変更'; ws['D2'].font=f_note
label(ws,'A4','■ 店舗規模',sub=True)
srow=[('打席数','=Assumptions!$B$5','#,##0'),('坪数','=Assumptions!$B$6','#,##0'),('会員上限','=Assumptions!$B$31','#,##0')]
r=5
for lb,val,fmt in srow:
    label(ws,f'A{r}',lb); calc(ws,f'B{r}',val,fmt,link=True); r+=1
label(ws,'A9','■ 資金',sub=True)
smap=[('総初期投資',f'={CAPEX_TOTAL}',YEN),('うち内装工事','=CAPEX!$B$5',YEN),('うち機材','=CAPEX!$B$9',YEN),
 ('出資金','=Assumptions!$B$28',YEN),('自己資金・融資','=Assumptions!$B$29',YEN),
 ('開業時運転資金','=Assumptions!$B$28+Assumptions!$B$29-'+CAPEX_TOTAL,YEN),
 ('最低現金残高','=MIN(Cash_Flow!$G$4:$G$40)',YEN),('資金ショート月数','=COUNTIF(Cash_Flow!$G$5:$G$40,"<0")','#,##0'),
 ('投資回収期間(月)','=IF(MIN(Cash_Flow!$I$5:$I$40)=9999,"36ヶ月内未回収",MIN(Cash_Flow!$I$5:$I$40))','General')]
r=10
for lb,val,fmt in smap:
    label(ws,f'A{r}',lb); calc(ws,f'B{r}',val,fmt,link=True); r+=1
label(ws,f'A{r+1}','■ 損益(年次)',sub=True); hdr(ws,r+2,['指標','Y1','Y2','Y3','36ヶ月累計']); hr=r+2
yl={'Y1':('4','15'),'Y2':('16','27'),'Y3':('28','39')}
rr=hr+1
for name,col in [('売上高','L'),('営業利益','V'),('税引後利益','Y')]:
    ws.cell(row=rr,column=1,value=name).font=Font(name=BASE); ws.cell(row=rr,column=1).border=border
    for i,yk in enumerate(['Y1','Y2','Y3']):
        s,e=yl[yk]; c=ws.cell(row=rr,column=2+i,value=f'=SUM(Monthly_PL!${col}${s}:${col}${e})'); c.font=f_link; c.number_format=YEN; c.border=border
    c=ws.cell(row=rr,column=5,value=f'=SUM(Monthly_PL!${col}$4:${col}$39)'); c.font=f_link; c.number_format=YEN; c.border=border
    rr+=1
label(ws,f'A{rr+1}','■ 主要指標',sub=True)
kmap=[('期末会員数(36ヶ月)','=Monthly_PL!$G$39','#,##0'),('損益分岐点会員数','=BreakEven!$B$8','#,##0'),('現状-BEP会員差','=Monthly_PL!$G$39-BreakEven!$B$8','#,##0')]
r=rr+2
for lb,val,fmt in kmap:
    label(ws,f'A{r}',lb); calc(ws,f'B{r}',val,fmt,link=True); r+=1
ws.column_dimensions['A'].width=26
for col in 'BCDE': ws.column_dimensions[col].width=15
ws['A'+str(r+2)]='注: 非居抜き前提の仮置き初版。物件条件確定で更新。'; ws['A'+str(r+2)].font=f_note

out=os.path.join(os.path.dirname(os.path.abspath(__file__)),'03_GOLFWING2_収支計画.xlsx')
wb.save(out); print('SAVED',out)
