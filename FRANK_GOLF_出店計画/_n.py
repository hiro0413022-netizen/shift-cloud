#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FRANK GOLF (GOLF WING 2号店) 収支計画モデル ビルドスクリプト
- 11シート / 楽観・標準・悲観・最悪の4シナリオ切替 / 36ヶ月 / 全数式連動
- 入力=青字＋黄背景、同一シート数式=黒字、他シート参照=緑字
実行: python 03_FRANK_GOLF収支計画_build.py
その後: python <xlsxスキル>/scripts/recalc.py 03_FRANK_GOLF収支計画.xlsx でエラー0を確認
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
import os

YEN = '#,##0;(#,##0);"-"'
PCT = '0.0%'
BASE = 'Arial'
f_title = Font(name=BASE, bold=True, size=13, color='1F3864')
f_h = Font(name=BASE, bold=True, color='FFFFFF')
f_sub = Font(name=BASE, bold=True, color='1F3864')
f_in = Font(name=BASE, color='0000FF')          # 入力
f_calc = Font(name=BASE, color='000000')         # 同一シート数式
f_link = Font(name=BASE, color='008000')         # 他シート参照
f_note = Font(name=BASE, italic=True, size=9, color='808080')
fill_in = PatternFill('solid', fgColor='FFFF00')
fill_h = PatternFill('solid', fgColor='1F3864')
fill_sub = PatternFill('solid', fgColor='D9E1F2')
fill_warn = PatternFill('solid', fgColor='FFC7CE')
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = Workbook()

def hdr(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start+i, value=h)
        c.font = f_h; c.fill = fill_h; c.alignment = center; c.border = border

def inp(ws, coord, val, fmt=YEN):
    c = ws[coord]; c.value = val; c.font = f_in; c.fill = fill_in
    c.number_format = fmt; c.border = border; return c

def calc(ws, coord, formula, fmt=YEN, link=False):
    c = ws[coord]; c.value = formula; c.font = f_link if link else f_calc
    c.number_format = fmt; c.border = border; return c

def label(ws, coord, text, sub=False):
    c = ws[coord]; c.value = text
    c.font = f_sub if sub else Font(name=BASE)
    if sub: c.fill = fill_sub
    c.alignment = left; return c

# ============================================================ Scenarios
ws = wb.active; ws.title = 'Scenarios'
ws['A1'] = 'シナリオ設定（4パターン比較）'; ws['A1'].font = f_title
ws['A2'] = '選択シナリオ →'; ws['A2'].font = f_sub
inp(ws, 'B2', '標準', fmt='General')
dv = DataValidation(type='list', formula1='"楽観,標準,悲観,最悪"', allow_blank=False)
ws.add_data_validation(dv); dv.add(ws['B2'])
ws['D2'] = 'B2で選ぶと全シート（PL/CF/損益分岐点）が自動更新'; ws['D2'].font = f_note

hdr(ws, 4, ['パラメータ', '楽観', '標準', '悲観', '最悪', '選択中(自動)', '単位/備考'])
# param rows 5..16
params = [
    # label, 楽観, 標準, 悲観, 最悪, fmt, note
    ('平均月会費(税抜)', 17000, 15000, 14000, 13000, YEN, '宝塚加重平均約18,800円を保守補正'),
    ('入会率(体験→入会)', 0.70, 0.60, 0.25, 0.18, PCT, '標準=体験15×60%想定。悲観/最悪は保守'),
    ('退会率(月次)', 0.03, 0.04, 0.06, 0.08, PCT, '宝塚実績3.5-3.9%。新店は保守'),
    ('月間体験数 Y1', 20, 15, 8, 5, '#,##0', '標準=15/月想定'),
    ('月間体験数 Y2', 22, 15, 12, 8, '#,##0', ''),
    ('月間体験数 Y3', 24, 15, 16, 10, '#,##0', ''),
    ('広告費 開業1-3ヶ月(月)', 80000, 100000, 150000, 200000, YEN, '開業直後の集客投資'),
    ('広告費 4ヶ月目以降(月)', 40000, 50000, 90000, 120000, YEN, '標準=5万/月'),
    ('人件費(月)', 500000, 580000, 650000, 750000, YEN, '無人化前提。最悪は増員'),
    ('物販ARPU(会員/月)', 3000, 2000, 1500, 1000, YEN, '会員1人あたり月物販'),
    ('開業時先行会員数', 40, 30, 5, 3, '#,##0', '標準=開業時30名想定'),
    ('追加CAPEX(初期不良/やり直し)', 0, 0, 300000, 800000, YEN, '悲観/最悪で発生'),
]
r0 = 5
for i, (lb, o, s, p, w, fmt, note) in enumerate(params):
    r = r0 + i
    label(ws, f'A{r}', lb)
    inp(ws, f'B{r}', o, fmt); inp(ws, f'C{r}', s, fmt); inp(ws, f'D{r}', p, fmt); inp(ws, f'E{r}', w, fmt)
    calc(ws, f'F{r}', f'=INDEX(B{r}:E{r},MATCH($B$2,$B$4:$E$4,0))', fmt)
    ws[f'G{r}'] = note; ws[f'G{r}'].font = f_note
SC_LAST = r0 + len(params) - 1  # 16
for col, wdt in zip('ABCDEFG', [24, 11, 11, 11, 11, 13, 34]):
    ws.column_dimensions[col].width = wdt

# Scenario active row map (F column): fee5 join6 churn7 tY1_8 tY2_9 tY3_10 adY1_11 adY23_12 labor13 arpu14 pre15 addcapex16
SF = {'fee':'F5','join':'F6','churn':'F7','tY1':'F8','tY2':'F9','tY3':'F10',
      'adY1':'F11','adY23':'F12','labor':'F13','arpu':'F14','pre':'F15','addcapex':'F16'}

# ============================================================ Assumptions
ws = wb.create_sheet('Assumptions')
ws['A1'] = '前提条件（入力の集約）'; ws['A1'].font = f_title
ws['A2'] = '黄=入力（変更可）／緑=他シート参照。他シートは原則ここを参照'; ws['A2'].font = f_note
label(ws, 'A4', '■ 基本前提', sub=True); ws['B4'].fill = fill_sub
rows_const = [
    ('会員上限(capacity)', 200, '#,##0'),          # B5
    ('入会金(税抜)', 30000, YEN),                    # B6
    ('体験単価(税抜)', 2000, YEN),                   # B7
    ('物販原価率', 0.65, PCT),                       # B8
    ('決済手数料率', 0.03, PCT),                     # B9
    ('決済適用率(カード等比率)', 0.90, PCT),          # B10
    ('減価償却年数(設備)', 7, '#,##0'),               # B11
    ('実効税率', 0.30, PCT),                         # B12
    ('家賃(月)', 400000, YEN),                       # B13
    ('光熱費(月)', 80000, YEN),                      # B14
    ('通信費(月)', 15000, YEN),                      # B15
    ('保険(月)', 12000, YEN),                        # B16
    ('システム費(月)', 10000, YEN),                  # B17 自社Genesis利用でインフラ/API実費のみ
    ('清掃費(月)', 20000, YEN),                      # B18
    ('修繕・設備更新積立(月)', 30000, YEN),          # B19
    ('消耗品費(月)', 15000, YEN),                    # B20
    ('税理士費用(月)', 0, YEN),                      # B21 YOZAN顧問税理士が対応
    ('予備費率(対その他固定費)', 0.03, PCT),          # B22
    ('保証金月数', 6, '#,##0'),                      # B23
    ('出資金', 15000000, YEN),                       # B24
    ('自己資金・融資', 0, YEN),                       # B25
]
r = 5
for lb, v, fmt in rows_const:
    label(ws, f'A{r}', lb); inp(ws, f'B{r}', v, fmt); r += 1
# r now 26
label(ws, 'A27', '■ 選択シナリオの反映値（自動）', sub=True)
active = [
    ('平均月会費(税抜)', SF['fee'], YEN),   # B28
    ('入会率', SF['join'], PCT),            # B29
    ('退会率(月)', SF['churn'], PCT),       # B30
    ('月間体験数 Y1', SF['tY1'], '#,##0'),  # B31
    ('月間体験数 Y2', SF['tY2'], '#,##0'),  # B32
    ('月間体験数 Y3', SF['tY3'], '#,##0'),  # B33
    ('広告費 開業1-3ヶ月(月)', SF['adY1'], YEN),     # B34
    ('広告費 4ヶ月目以降(月)', SF['adY23'], YEN),    # B35
    ('人件費(月)', SF['labor'], YEN),       # B36
    ('物販ARPU(月)', SF['arpu'], YEN),      # B37
    ('開業時先行会員数', SF['pre'], '#,##0'),# B38
    ('追加CAPEX', SF['addcapex'], YEN),     # B39
]
r = 28
for lb, ref, fmt in active:
    label(ws, f'A{r}', lb); calc(ws, f'B{r}', f'=Scenarios!{ref}', fmt, link=True); r += 1
# r now 40
label(ws, 'A41', '■ その他固定費の合計（自動）', sub=True)
label(ws, 'A42', 'その他固定費小計(家賃/人件費/広告除く)')
calc(ws, 'B42', '=B14+B15+B16+B17+B18+B19+B20+B21', YEN)
label(ws, 'A43', '予備費'); calc(ws, 'B43', '=B42*B22', YEN)
label(ws, 'A44', 'その他固定費 計'); calc(ws, 'B44', '=B42+B43', YEN)
label(ws, 'A46', '■ 前受け（月会費先取り）', sub=True)
label(ws, 'A47', '前受けON(1=有効,0=無効)'); inp(ws, 'B47', 1, '#,##0')
ws['C47'] = '3ヶ月前取り等の運転資本効果をCFに反映'; ws['C47'].font = f_note
label(ws, 'A48', '前受けフロート係数(月分)'); inp(ws, 'B48', 1.5, '0.0')
ws['C48'] = '3ヶ月前取り≒1.5。会員増加分×係数×会費が現金前倒し'; ws['C48'].font = f_note
label(ws, 'A49', '■ 後日投資（防犯・セキュリティ）', sub=True)
label(ws, 'A50', '防犯・セキュリティ投資額'); inp(ws, 'B50', 400000, YEN)
label(ws, 'A51', '実施月(開業何ヶ月後)'); inp(ws, 'B51', 6, '#,##0')
ws['C50'] = '24時間無人化。開業後に実施しCAPEXを後ろ倒し'; ws['C50'].font = f_note
ws.column_dimensions['A'].width = 34; ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 40
# ref map
A = {'cap':'Assumptions!$B$5','join_fee':'Assumptions!$B$6','trial_p':'Assumptions!$B$7',
     'cogs':'Assumptions!$B$8','pay':'Assumptions!$B$9','payr':'Assumptions!$B$10',
     'dep_y':'Assumptions!$B$11','tax':'Assumptions!$B$12','rent':'Assumptions!$B$13',
     'fee':'Assumptions!$B$28','join':'Assumptions!$B$29','churn':'Assumptions!$B$30',
     'tY1':'Assumptions!$B$31','tY2':'Assumptions!$B$32','tY3':'Assumptions!$B$33',
     'adY1':'Assumptions!$B$34','adY23':'Assumptions!$B$35','labor':'Assumptions!$B$36',
     'arpu':'Assumptions!$B$37','pre':'Assumptions!$B$38','addcapex':'Assumptions!$B$39',
     'otherfix':'Assumptions!$B$44'}

# ============================================================ CAPEX
ws = wb.create_sheet('CAPEX')
ws['A1'] = '初期投資（CAPEX）内訳'; ws['A1'].font = f_title
ws['A2'] = '※既存楽観案(約10,340千円)は保証金・予備費・開業前費用が未計上の疑い。本表は保守計上。'; ws['A2'].font = f_note
hdr(ws, 4, ['項目', '金額', '備考'])
capex = [
    ('内装工事(居抜き)', 1800000, '居抜き活用で圧縮。造作・打席ブース中心'),
    ('電気工事', 800000, '分電盤・打席電源・照明'),
    ('空調・換気・設備工事', 1000000, '打席発熱対応で増設'),
    ('防音・安全対策', 800000, '防音・防球・防護'),
    ('シミュレーター・機材(2打席)', 2500000, '機種で50万〜250万/席。中位想定'),
    ('PC・ネットワーク・AV', 500000, ''),
    ('什器・備品', 500000, ''),
    ('看板・サイン', 400000, ''),
    ('決済・POS環境', 150000, ''),
    ('HP制作・初期ブランディング', 300000, 'YOZAN WEB。既存テンプレ流用'),
    ('開業前費用(採用・研修・内覧)', 300000, ''),
]  # 防犯・セキュリティは初期投資外（開業N ヶ月後に実施。Assumptions B49/B50）
r = 5
for lb, v, note in capex:
    label(ws, f'A{r}', lb); inp(ws, f'B{r}', v)
    ws[f'C{r}'] = note; ws[f'C{r}'].font = f_note; r += 1
last = r - 1  # 16
label(ws, f'A{r}', '設備小計', sub=True); calc(ws, f'B{r}', f'=SUM(B5:B{last})'); SUB=r; r+=1
label(ws, f'A{r}', '予備費(10%)', sub=True); calc(ws, f'B{r}', f'=B{SUB}*0.1'); PRE=r; r+=1
label(ws, f'A{r}', 'CAPEX計(減価償却対象)', sub=True); calc(ws, f'B{r}', f'=B{SUB}+B{PRE}'); DEPB=r; r+=1
label(ws, f'A{r}', '保証金・敷金', sub=True); calc(ws, f'B{r}', '=Assumptions!$B$13*Assumptions!$B$23', YEN, link=True)
DEPO=r; r+=1
label(ws, f'A{r}', '追加CAPEX(シナリオ)', sub=True); calc(ws, f'B{r}', f'={A["addcapex"]}', YEN, link=True); ADD=r; r+=1
label(ws, f'A{r}', '総初期投資', sub=True); calc(ws, f'B{r}', f'=B{DEPB}+B{DEPO}+B{ADD}'); TOT=r
ws[f'B{TOT}'].font = Font(name=BASE, bold=True, color='C00000')
CAPEX_DEPBASE = f'CAPEX!$B${DEPB}'; CAPEX_ADD = f'CAPEX!$B${ADD}'; CAPEX_TOTAL = f'CAPEX!$B${TOT}'
ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 14; ws.column_dimensions['C'].width = 34

# ============================================================ OPEX
ws = wb.create_sheet('OPEX')
ws['A1'] = '月次コスト内訳（OPEX）'; ws['A1'].font = f_title
hdr(ws, 3, ['区分', '項目', '金額(月)', '種別', '備考'])
opex = [
    ('固定費', '家賃', '=Assumptions!$B$13', '固定', ''),
    ('固定費', '人件費', '=Assumptions!$B$36', '固定(シナリオ)', '最悪は増員'),
    ('固定費', '広告費(開業1-3ヶ月)', '=Assumptions!$B$34', '固定(期間変動)', '開業直後'),
    ('固定費', '広告費(4ヶ月〜)', '=Assumptions!$B$35', '固定(期間変動)', '4ヶ月目以降'),
    ('固定費', '光熱費', '=Assumptions!$B$14', '固定', ''),
    ('固定費', '通信費', '=Assumptions!$B$15', '固定', ''),
    ('固定費', '保険', '=Assumptions!$B$16', '固定', ''),
    ('固定費', 'システム費', '=Assumptions!$B$17', '固定', '予約/会員/機材月額'),
    ('固定費', '清掃費', '=Assumptions!$B$18', '固定', ''),
    ('固定費', '修繕・設備更新積立', '=Assumptions!$B$19', '固定', ''),
    ('固定費', '消耗品費', '=Assumptions!$B$20', '固定', ''),
    ('固定費', '税理士費用', '=Assumptions!$B$21', '固定', ''),
    ('固定費', '予備費', '=Assumptions!$B$43', '固定', '対その他固定費3%'),
    ('変動費', '決済手数料', '売上×3%×適用率', '変動', 'Monthly_PLで計算'),
    ('変動費', '物販原価', '物販売上×65%', '変動', 'Monthly_PLで計算'),
]
r = 4
for seg, it, val, typ, note in opex:
    ws[f'A{r}'] = seg; ws[f'B{r}'] = it
    if isinstance(val, str) and val.startswith('='):
        calc(ws, f'C{r}', val, YEN, link=True)
    else:
        ws[f'C{r}'] = val; ws[f'C{r}'].font = f_note
    ws[f'D{r}'] = typ; ws[f'E{r}'] = note; ws[f'E{r}'].font = f_note
    for col in 'ABDE': ws[f'{col}{r}'].border = border
    r += 1
label(ws, f'A{r+1}', '固定費計(Y1,広告Y1)', sub=True)
calc(ws, f'C{r+1}', '=C4+C5+C6+C8+C9+C10+C11+C12+C13+C14+C15+C16', YEN)
label(ws, f'A{r+2}', '固定費計(Y2-3,広告Y2-3)', sub=True)
calc(ws, f'C{r+2}', '=C4+C5+C7+C8+C9+C10+C11+C12+C13+C14+C15+C16', YEN)
for col, wdt in zip('ABCDE', [10, 22, 14, 14, 26]): ws.column_dimensions[col].width = wdt

# ============================================================ Monthly_PL
ws = wb.create_sheet('Monthly_PL')
ws['A1'] = '月次損益計画（36ヶ月）'; ws['A1'].font = f_title
ws['A2'] = '会員動態→売上→変動費→固定費→営業利益。選択シナリオで自動更新。'; ws['A2'].font = f_note
cols = ['月','年','期首会員','体験数','入会数','退会数','期末会員','月会費売上','入会金売上',
        '体験売上','物販売上','売上合計','物販原価','決済手数料','変動費計','家賃','人件費',
        '広告費','その他固定費','固定費計','減価償却','営業利益','税前利益','法人税等','税引後利益']
hdr(ws, 3, cols)
HR = 3
for m in range(1, 37):
    r = HR + m  # month1=row4 ... month36=row39
    ws[f'A{r}'] = m
    ws[f'B{r}'] = f'=ROUNDUP(A{r}/12,0)'
    # 期首
    if m == 1:
        ws[f'C{r}'] = f'={A["pre"]}'; ws[f'C{r}'].font=f_link
    else:
        ws[f'C{r}'] = f'=G{r-1}'; ws[f'C{r}'].font=f_calc
    # 体験数 = 年別体験 × ランプ(MIN(1,月/6))
    ws[f'D{r}'] = (f'=ROUND(IF(B{r}=1,{A["tY1"]},IF(B{r}=2,{A["tY2"]},{A["tY3"]}))'
                   f'*MIN(1,A{r}/6),0)')
    # 退会
    ws[f'F{r}'] = f'=ROUND(C{r}*{A["churn"]},0)'
    # 入会 = MIN(体験×入会率, capacity-(期首-退会)), >=0
    ws[f'E{r}'] = f'=MAX(0,MIN(ROUND(D{r}*{A["join"]},0),{A["cap"]}-(C{r}-F{r})))'
    # 期末
    ws[f'G{r}'] = f'=MAX(0,MIN({A["cap"]},C{r}+E{r}-F{r}))'
    # 売上
    ws[f'H{r}'] = f'=C{r}*{A["fee"]}'
    ws[f'I{r}'] = f'=E{r}*{A["join_fee"]}'
    ws[f'J{r}'] = f'=D{r}*{A["trial_p"]}'
    ws[f'K{r}'] = f'=C{r}*{A["arpu"]}'
    ws[f'L{r}'] = f'=SUM(H{r}:K{r})'
    # 変動費
    ws[f'M{r}'] = f'=K{r}*{A["cogs"]}'
    ws[f'N{r}'] = f'=L{r}*{A["pay"]}*{A["payr"]}'
    ws[f'O{r}'] = f'=M{r}+N{r}'
    # 固定費
    ws[f'P{r}'] = f'={A["rent"]}'
    ws[f'Q{r}'] = f'={A["labor"]}'
    ws[f'R{r}'] = f'=IF(A{r}<=3,{A["adY1"]},{A["adY23"]})'
    ws[f'S{r}'] = f'={A["otherfix"]}'
    ws[f'T{r}'] = f'=P{r}+Q{r}+R{r}+S{r}'
    # 減価償却
    ws[f'U{r}'] = f'=({CAPEX_DEPBASE}+{CAPEX_ADD})/({A["dep_y"]}*12)+IF(A{r}>=Assumptions!$B$51,Assumptions!$B$50/({A["dep_y"]}*12),0)'
    # 利益
    ws[f'V{r}'] = f'=L{r}-O{r}-T{r}-U{r}'
    ws[f'W{r}'] = f'=V{r}'
    ws[f'X{r}'] = f'=MAX(0,W{r})*{A["tax"]}'
    ws[f'Y{r}'] = f'=W{r}-X{r}'
    for ci in range(1, 26):
        c = ws.cell(row=r, column=ci); c.border = border
        col = c.column_letter
        if col in ('A','B','C','D','E','F','G'):
            c.number_format = '#,##0'
        else:
            c.number_format = YEN
        # font: green if references other sheet
        if col in ('C','D','E','F','G','H','I','J','K','M','N','P','Q','R','S','U') and m >= 1:
            pass
        c.font = f_link if (isinstance(c.value,str) and 'Assumptions!' in str(c.value) or 'CAPEX!' in str(c.value)) else f_calc
for col in 'ABCDEFG': ws.column_dimensions[col].width = 8
for col in ['H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y']:
    ws.column_dimensions[col].width = 11
ws.freeze_panes = 'C4'
PL_L='Monthly_PL!$L'; PL_V='Monthly_PL!$V'; PL_Y='Monthly_PL!$Y'; PL_G='Monthly_PL!$G'; PL_U='Monthly_PL!$U'

# ============================================================ Cash_Flow
ws = wb.create_sheet('Cash_Flow')
ws['A1'] = '月次キャッシュフロー'; ws['A1'].font = f_title
ws['A2'] = '期末現金がマイナス＝資金ショート（赤表示）。営業CF=税引後利益+減価償却。'; ws['A2'].font = f_note
hdr(ws, 3, ['月','営業CF','投資CF','財務CF','前受けCF(運転資本)','純増減','期末現金','累計営業CF','回収判定(内部)'])
# 列: A月 B営業CF C投資CF D財務CF E前受けCF F純増減 G期末現金 H累計営業CF I回収判定
# month0 row4
ws['A4'] = 0
ws['B4'] = 0
ws['C4'] = f'=-{CAPEX_TOTAL}'; ws['C4'].font=f_link
ws['D4'] = '=Assumptions!$B$24+Assumptions!$B$25'; ws['D4'].font=f_link
ws['E4'] = 0
ws['F4'] = '=B4+C4+D4+E4'
ws['G4'] = '=F4'
ws['H4'] = 0
ws['I4'] = 9999
for col in 'ABCDEFGHI':
    ws[f'{col}4'].border=border; ws[f'{col}4'].number_format=YEN
for m in range(1, 37):
    r = 4 + m  # month1 row5 ... month36 row40
    plr = 3 + m  # Monthly_PL row for month m
    prev = '0' if m == 1 else f'Monthly_PL!$G${plr-1}'
    ws[f'A{r}'] = m
    ws[f'B{r}'] = f'={PL_Y}{plr}+{PL_U}{plr}'; ws[f'B{r}'].font=f_link
    ws[f'C{r}'] = f'=IF(A{r}=Assumptions!$B$51,-Assumptions!$B$50,0)'; ws[f'C{r}'].font=f_link
    ws[f'D{r}'] = 0
    ws[f'E{r}'] = f'=Assumptions!$B$47*Assumptions!$B$48*Assumptions!$B$28*(Monthly_PL!$G${plr}-{prev})'
    ws[f'E{r}'].font=f_link
    ws[f'F{r}'] = f'=B{r}+C{r}+D{r}+E{r}'
    ws[f'G{r}'] = f'=G{r-1}+F{r}'
    ws[f'H{r}'] = f'=SUM($B$5:B{r})'
    ws[f'I{r}'] = f'=IF(H{r}>={CAPEX_TOTAL},A{r},9999)'
    for col in 'ABCDEFGHI':
        ws[f'{col}{r}'].border=border; ws[f'{col}{r}'].number_format=YEN
    ws[f'A{r}'].number_format='#,##0'
ws.conditional_formatting.add('G5:G40', CellIsRule(operator='lessThan', formula=['0'], fill=fill_warn, font=Font(name=BASE,color='9C0006')))
for col, wdt in zip('ABCDEFGHI',[8,13,11,11,15,13,14,14,12]): ws.column_dimensions[col].width=wdt
ws.freeze_panes='B4'
CF_F='Cash_Flow!$F'; CF_G='Cash_Flow!$G'; CF_A='Cash_Flow!$A'

# ============================================================ BreakEven
ws = wb.create_sheet('BreakEven')
ws['A1'] = '損益分岐点分析'; ws['A1'].font = f_title
rows = [
    ('月固定費(定常/Y2-3)', '=Assumptions!$B$13+Assumptions!$B$36+Assumptions!$B$35+Assumptions!$B$44', YEN, True),
    ('平均月会費(税抜)', '=Assumptions!$B$28', YEN, True),
    ('物販ARPU', '=Assumptions!$B$37', YEN, True),
    ('物販原価率', '=Assumptions!$B$8', PCT, True),
    ('決済実効率(手数料率×適用率)', '=Assumptions!$B$9*Assumptions!$B$10', PCT, True),
    ('会員あたり月次貢献', '=B3*(1-B6)+B4*(1-B5)-B4*B6', YEN, False),
    ('損益分岐点 会員数', '=ROUNDUP(B2/B7,0)', '#,##0', False),
    ('必要月商(概算)', '=B8*(B3+B4)', YEN, False),
    ('現状会員(36ヶ月)', '=Monthly_PL!$G$39', '#,##0', True),
    ('BEP-現状会員 差', '=B8-B10', '#,##0', False),
    ('BEP維持に必要な月間入会(退会補填)', '=ROUNDUP(B8*Assumptions!$B$30,0)', '#,##0', False),
    ('必要月間体験数', '=ROUNDUP(B11/Assumptions!$B$29,0)', '#,##0', False),
]
r = 2
for lb, val, fmt, link in rows:
    label(ws, f'A{r}', lb); calc(ws, f'B{r}', val, fmt, link=link); r += 1
ws.column_dimensions['A'].width=32; ws.column_dimensions['B'].width=14
BE_BEP='BreakEven!$B$8'

# ============================================================ Sensitivity
ws = wb.create_sheet('Sensitivity')
ws['A1'] = '感応度分析（年間営業利益への影響・一次近似）'; ws['A1'].font = f_title
ws['A2'] = '精緻な再計算はAssumptionsを変更しSummaryで確認。下表は目安。'; ws['A2'].font = f_note
calc(ws, 'B4', '=SUM(Monthly_PL!$V$28:$V$39)', YEN, link=True)
label(ws, 'A4', '基準:年間営業利益(Y3/25-36月)', sub=True)
hdr(ws, 6, ['変動要因', '変化幅', '年間営業利益への影響(近似)'])
sens = [
    ('会員数', '+10%', '=0.1*Monthly_PL!$G$39*BreakEven!$B$7*12'),
    ('会員数', '-10%', '=-0.1*Monthly_PL!$G$39*BreakEven!$B$7*12'),
    ('平均月会費', '+1,000円', '=Monthly_PL!$G$39*1000*(1-Assumptions!$B$9*Assumptions!$B$10)*12'),
    ('平均月会費', '-1,000円', '=-Monthly_PL!$G$39*1000*(1-Assumptions!$B$9*Assumptions!$B$10)*12'),
    ('退会率', '+2pt', '=-Monthly_PL!$G$39*0.02*BreakEven!$B$7*12'),
    ('広告費', '+20%', '=-Assumptions!$B$35*0.2*12'),
    ('人件費', '+10%', '=-Assumptions!$B$36*0.1*12'),
    ('家賃', '+10%', '=-Assumptions!$B$13*0.1*12'),
]
r = 7
for f_, d_, v_ in sens:
    ws[f'A{r}'] = f_; ws[f'B{r}'] = d_; calc(ws, f'C{r}', v_, YEN, link=True)
    for col in 'AB': ws[f'{col}{r}'].border=border
    r += 1
for col, wdt in zip('ABC',[16,14,28]): ws.column_dimensions[col].width=wdt

# ============================================================ Risks
ws = wb.create_sheet('Risks')
ws['A1'] = 'リスク一覧（影響度・発生可能性・対策）'; ws['A1'].font = f_title
hdr(ws, 3, ['分類','リスク','影響度','発生可能性','対策'])
risks = [
    ('出店','良物件が出ない/賃料高','高','中','探索チャネル複数化・居抜き優先・条件緩和'),
    ('集客','初年度の会員立ち上がり遅延','高','高','先行入会・紹介・体験導線一本化・広告前倒し'),
    ('集客','入会率が想定を下回る','高','中','体験の質・接客標準化・フォロー導線'),
    ('運営','退会率上昇(定着不足)','高','中','予約取りやすさ改善・レッスン価値・休会提案'),
    ('資金繰り','初期投資超過で運転資金枯渇','高','高','CAPEX圧縮(段階投資)・追加融資枠・出資増額'),
    ('資金繰り','資金ショート月の発生','高','中','固定費の変動費化・支払サイト調整・与信枠'),
    ('人材','採用難/教育不足','中','中','無人化比率UP・宝塚応援・動画教材'),
    ('設備','シミュレーター故障/納期','中','中','保守契約・予備部材・中位機で先行開業'),
    ('法務','ブランド/名称使用の条件相違','高','中','ファインと書面合意(確認事項参照)'),
    ('契約','賃貸/原状回復/中途解約の負担','中','中','契約前の条件精査・弁護士確認'),
    ('近隣','騒音クレーム','中','低','事前騒音測定・防音余裕設計'),
    ('ブランド','運営品質のばらつきでGOLF WING毀損','高','中','SOP整備・KPI監視・是正フロー'),
]
r = 4
for cat, rk, imp, prob, act in risks:
    for i, v in enumerate([cat, rk, imp, prob, act]):
        c = ws.cell(row=r, column=1+i, value=v); c.border=border; c.alignment=left
    r += 1
for col, wdt in zip('ABCDE',[10,30,8,10,40]): ws.column_dimensions[col].width=wdt

# ============================================================ Investor_QA
ws = wb.create_sheet('Investor_QA')
ws['A1'] = '想定問答（ファイン/金融機関/出資者）'; ws['A1'].font = f_title
hdr(ws, 3, ['想定質問','回答案'])
qa = [
    ('売上前提は過大では?','宝塚の加重平均会費18,800円を15,000円へ保守補正。入会率も純転換のみ採用。悲観/最悪で退会率6-8%を検証。'),
    ('初期投資は足りるか?','既存楽観案は保証金・予備費・開業前費用が未計上。本モデルで保守計上した結果、総初期投資は出資1,500万を上回り得るため、CAPEX圧縮or追加調達を明示。'),
    ('資金ショートは?','Cash_Flowで月末現金を可視化。標準では回避、悲観/最悪では発生月を赤表示。対策(段階投資/融資枠)を用意。'),
    ('投資回収は?','Summaryに投資回収月を表示。標準で概ね2-3年目、悲観で長期化。'),
    ('集客が失敗したら?','最悪シナリオで初期集客失敗・退会増・広告増・人件費増・修繕発生を同時反映し耐性を確認。'),
    ('損益分岐点は?','BreakEvenで会員数・必要月商・必要体験数を提示。定常固定費÷会員貢献。'),
    ('ファインとの関係は?','名称・ブランド・システム利用の範囲と責任分担を書面合意(確認事項一覧)。'),
    ('人が採れなかったら?','無人運営モデルで必要人員を圧縮。宝塚からの応援と業務委託で補完。'),
    ('競合が増えたら?','会員制+レッスン+無人+システム化で差別化。KPI監視で早期対応。'),
    ('撤退基準は?','一定期間の会員数/現金残高が基準割れで段階縮小・撤退を事前定義(要協議)。'),
]
r = 4
for q, a in qa:
    ws.cell(row=r, column=1, value=q).border=border
    ws.cell(row=r, column=1).alignment=left
    ws.cell(row=r, column=2, value=a).border=border
    ws.cell(row=r, column=2).alignment=left
    r += 1
ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=70

# ============================================================ Summary (先頭へ)
ws = wb.create_sheet('Summary', 0)
ws['A1'] = 'FRANK GOLF (GOLF WING 2号店) 収支サマリー'; ws['A1'].font = f_title
ws['A2'] = '選択シナリオ:'; ws['A2'].font=f_sub
calc(ws, 'B2', '=Scenarios!$B$2', 'General', link=True); ws['B2'].font=Font(name=BASE,bold=True,color='C00000')
ws['D2'] = 'シナリオはScenarios!B2で変更'; ws['D2'].font=f_note
label(ws, 'A4', '■ 資金', sub=True)
smap = [
    ('総初期投資', f'={CAPEX_TOTAL}', YEN),
    ('出資金', '=Assumptions!$B$24', YEN),
    ('自己資金・融資', '=Assumptions!$B$25', YEN),
    ('開業時運転資金(手元)', '=Assumptions!$B$24+Assumptions!$B$25-'+CAPEX_TOTAL, YEN),
    ('最低現金残高(0-36月)', '=MIN(Cash_Flow!$G$4:$G$40)', YEN),
    ('資金ショート月数', '=COUNTIF(Cash_Flow!$G$5:$G$40,"<0")', '#,##0'),
    ('投資回収期間(月)', '=IF(MIN(Cash_Flow!$I$5:$I$40)=9999,"36ヶ月内未回収",MIN(Cash_Flow!$I$5:$I$40))', 'General'),
]
r = 5
for lb, val, fmt in smap:
    label(ws, f'A{r}', lb); calc(ws, f'B{r}', val, fmt, link=True); r += 1
label(ws, f'A{r+1}', '■ 損益(年次)', sub=True)
hdr(ws, r+2, ['指標','Y1','Y2','Y3','36ヶ月累計'])
hr = r+2
yl = {'Y1':('4','15'),'Y2':('16','27'),'Y3':('28','39')}
metrics = [
    ('売上高', 'L'),
    ('営業利益', 'V'),
    ('税引後利益', 'Y'),
]
rr = hr+1
for name, col in metrics:
    ws.cell(row=rr, column=1, value=name).font=Font(name=BASE)
    ws.cell(row=rr, column=1).border=border
    for i,(yk) in enumerate(['Y1','Y2','Y3']):
        s,e = yl[yk]
        c = ws.cell(row=rr, column=2+i, value=f'=SUM(Monthly_PL!${col}${s}:${col}${e})')
        c.font=f_link; c.number_format=YEN; c.border=border
    c = ws.cell(row=rr, column=5, value=f'=SUM(Monthly_PL!${col}$4:${col}$39)')
    c.font=f_link; c.number_format=YEN; c.border=border
    rr += 1
label(ws, f'A{rr+1}', '■ 主要指標', sub=True)
kmap = [
    ('期末会員数(36ヶ月)', '=Monthly_PL!$G$39', '#,##0'),
    ('損益分岐点会員数', '=BreakEven!$B$8', '#,##0'),
    ('現状-BEP会員差', '=Monthly_PL!$G$39-BreakEven!$B$8', '#,##0'),
]
r = rr+2
for lb, val, fmt in kmap:
    label(ws, f'A{r}', lb); calc(ws, f'B{r}', val, fmt, link=True); r += 1
ws.column_dimensions['A'].width=26
for col in 'BCDE': ws.column_dimensions[col].width=15
ws['A'+str(r+2)] = '注: 数値は仮置き前提に基づく初版。実データ確定で更新。'; ws['A'+str(r+2)].font=f_note

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), '03_FRANK_GOLF収支計画.xlsx')
wb.save(out)
print('SAVED', out)
